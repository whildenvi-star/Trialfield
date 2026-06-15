"""Tests for all seven output files.

Uses a lightweight synthetic trial (3 numeric treatments, 2 reps) for most tests
so they run quickly without the full Schultz geometry stack.  A separate section
re-uses the Schultz regression fixture for integration-level checks.
"""

from __future__ import annotations

import csv
import math
import os
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pytest
import shapefile

from trialfield_core.geometry.placement import RepBlock
from trialfield_core.geometry.plots import PlotRecord, generate_plots, randomize_treatments
from trialfield_core.geometry.uv_frame import UVFrame
from trialfield_core.io.abline import read_ab_line, write_ab_line
from trialfield_core.models.geometry_inputs import ABLine
from trialfield_core.models.soil import MajoritySoilZone, SSURGOComponent
from trialfield_core.models.trial_design import Treatment, TrialDesign, TrialType
from trialfield_core.outputs.csv_export import write_plots_csv
from trialfield_core.outputs.kml import write_kml
from trialfield_core.outputs.map_render import write_map
from trialfield_core.outputs.sample_pins import write_sample_pins
from trialfield_core.outputs.shapefile import write_rx_shapefile
from trialfield_core.outputs.summary import write_summary

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

_AB = ABLine(a_lon=-89.10575, a_lat=42.56241, b_lon=-89.11105, b_lat=42.56250, bearing_deg=271.4)

_NUMERIC_TREATMENTS = [
    Treatment(label="0", value=0.0, unit="lb N/ac"),
    Treatment(label="100", value=100.0, unit="lb N/ac"),
    Treatment(label="200", value=200.0, unit="lb N/ac"),
]

_CATEGORICAL_TREATMENTS = [
    Treatment(label="Tank Mix A"),
    Treatment(label="Tank Mix B"),
    Treatment(label="Untreated"),
]

_BLOCKS_2REP = [
    RepBlock(rep=1, label="W", u_west=-400.0, u_east=0.0, v_south=300.0, v_north=480.0, strip_order=[0, 1, 2]),
    RepBlock(rep=2, label="E", u_west=0.0, u_east=400.0, v_south=300.0, v_north=480.0, strip_order=[0, 1, 2]),
]

_TREATMENT_ASSIGN = {1: [0, 1, 2], 2: [2, 0, 1]}


@pytest.fixture()
def frame() -> UVFrame:
    return UVFrame.from_ab_wgs84(_AB.a_lon, _AB.a_lat, _AB.b_lon, _AB.b_lat)


@pytest.fixture()
def numeric_trial() -> TrialDesign:
    return TrialDesign(
        name="N-rate synthetic",
        trial_type=TrialType.fertility,
        treatments=_NUMERIC_TREATMENTS,
        reps=2,
    )


@pytest.fixture()
def categorical_trial() -> TrialDesign:
    return TrialDesign(
        name="Tank mix synthetic",
        trial_type=TrialType.spray,
        treatments=_CATEGORICAL_TREATMENTS,
        reps=2,
        plot_length_ft=300.0,
    )


@pytest.fixture()
def numeric_plots(frame: UVFrame, numeric_trial: TrialDesign) -> list[PlotRecord]:
    return generate_plots(
        blocks=_BLOCKS_2REP,
        treatment_assignments=_TREATMENT_ASSIGN,
        treatments=numeric_trial.treatments,
        swath_width_ft=60.0,
        frame=frame,
    )


@pytest.fixture()
def categorical_plots(frame: UVFrame, categorical_trial: TrialDesign) -> list[PlotRecord]:
    return generate_plots(
        blocks=_BLOCKS_2REP,
        treatment_assignments=_TREATMENT_ASSIGN,
        treatments=categorical_trial.treatments,
        swath_width_ft=60.0,
        frame=frame,
    )


@pytest.fixture()
def soil_ok() -> MajoritySoilZone:
    return MajoritySoilZone(
        wkt="POLYGON((-89.12 42.56,-89.10 42.56,-89.10 42.57,-89.12 42.57,-89.12 42.56))",
        components=[SSURGOComponent(mukey="1", compname="Drummer silty clay loam", comppct_r=80.0, hydgrp="A")],
        source="SDA",
    )


@pytest.fixture()
def soil_unavailable() -> MajoritySoilZone:
    return MajoritySoilZone(wkt="", components=[], source="unavailable", note="Timed out")


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


def test_csv_created(tmp_path, numeric_plots, numeric_trial):
    out = write_plots_csv(numeric_plots, numeric_trial, "synthetic", tmp_path)
    assert out.exists()
    assert out.suffix == ".csv"


def test_csv_row_count(tmp_path, numeric_plots, numeric_trial):
    out = write_plots_csv(numeric_plots, numeric_trial, "synthetic", tmp_path)
    with out.open() as f:
        rows = list(csv.reader(f))
    assert len(rows) == len(numeric_plots) + 1  # header + data


def test_csv_numeric_rate_column(tmp_path, numeric_plots, numeric_trial):
    out = write_plots_csv(numeric_plots, numeric_trial, "synthetic", tmp_path)
    with out.open() as f:
        header = next(csv.reader(f))
    # Unit "lb N/ac" → "lbN_ac"
    assert any("Rate" in col for col in header)
    assert "Rate_lbN_ac" in header


def test_csv_categorical_treatment_column(tmp_path, categorical_plots, categorical_trial):
    out = write_plots_csv(categorical_plots, categorical_trial, "synthetic", tmp_path)
    with out.open() as f:
        header = next(csv.reader(f))
    assert "Treatment_Label" in header


def test_csv_has_latlon_columns(tmp_path, numeric_plots, numeric_trial):
    out = write_plots_csv(numeric_plots, numeric_trial, "synthetic", tmp_path)
    with out.open() as f:
        header = next(csv.reader(f))
    for col in ("SW_Lat", "SW_Lon", "NE_Lat", "NE_Lon"):
        assert col in header, f"Missing column: {col}"


def test_csv_values_are_numbers(tmp_path, numeric_plots, numeric_trial):
    out = write_plots_csv(numeric_plots, numeric_trial, "synthetic", tmp_path)
    with out.open() as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        float(row["SW_Lat"])
        float(row["SW_Lon"])
        float(row["Acres"])


# ---------------------------------------------------------------------------
# FieldView Rx shapefile (numeric)
# ---------------------------------------------------------------------------


def test_rx_shapefile_created(tmp_path, numeric_plots, numeric_trial):
    out = write_rx_shapefile(numeric_plots, "synthetic", tmp_path)
    assert out is not None
    assert out.exists()
    assert out.suffix == ".zip"


def test_rx_shapefile_record_count(tmp_path, numeric_plots, numeric_trial):
    out = write_rx_shapefile(numeric_plots, "synthetic", tmp_path)
    with zipfile.ZipFile(out) as z:
        stem = [n for n in z.namelist() if n.endswith(".shp")][0].replace(".shp", "")
        with tempfile.TemporaryDirectory() as td:
            z.extractall(td)
            sf = shapefile.Reader(os.path.join(td, stem))
            assert len(sf.records()) == len(numeric_plots)


def test_rx_shapefile_has_rate_field(tmp_path, numeric_plots, numeric_trial):
    out = write_rx_shapefile(numeric_plots, "synthetic", tmp_path)
    with zipfile.ZipFile(out) as z:
        stem = [n for n in z.namelist() if n.endswith(".shp")][0].replace(".shp", "")
        with tempfile.TemporaryDirectory() as td:
            z.extractall(td)
            sf = shapefile.Reader(os.path.join(td, stem))
            field_names = [f[0] for f in sf.fields[1:]]
            assert "Rate" in field_names
            assert "Plot_ID" in field_names


def test_rx_shapefile_has_prj(tmp_path, numeric_plots):
    out = write_rx_shapefile(numeric_plots, "synthetic", tmp_path)
    with zipfile.ZipFile(out) as z:
        assert any(n.endswith(".prj") for n in z.namelist())


# ---------------------------------------------------------------------------
# Categorical trial — no Rx shapefile
# ---------------------------------------------------------------------------


def test_categorical_rx_shapefile_is_none(tmp_path, categorical_plots):
    result = write_rx_shapefile(categorical_plots, "synthetic", tmp_path)
    assert result is None


def test_categorical_no_zip_written(tmp_path, categorical_plots):
    write_rx_shapefile(categorical_plots, "synthetic", tmp_path)
    zips = list(tmp_path.glob("*.zip"))
    assert len(zips) == 0


# ---------------------------------------------------------------------------
# AB line shapefile round-trip
# ---------------------------------------------------------------------------


def test_ab_line_roundtrip(tmp_path):
    ab = ABLine(a_lon=-89.10575128, a_lat=42.56240918, b_lon=-89.11104903, b_lat=42.56250344, bearing_deg=271.38)
    zip_path = tmp_path / "ab_line.zip"
    write_ab_line(ab, zip_path)
    assert zip_path.exists()

    ab2 = read_ab_line(zip_path)
    assert abs(ab2.a_lon - ab.a_lon) < 1e-6
    assert abs(ab2.a_lat - ab.a_lat) < 1e-6
    assert abs(ab2.b_lon - ab.b_lon) < 1e-6
    assert abs(ab2.b_lat - ab.b_lat) < 1e-6


def test_ab_line_roundtrip_geometry_preserved(tmp_path):
    """The A-B point geometry survives write→read to within 1e-7 degrees (~1 cm)."""
    ab = ABLine(a_lon=-89.10575128, a_lat=42.56240918, b_lon=-89.11104903, b_lat=42.56250344)
    zip_path = tmp_path / "ab_rt.zip"
    write_ab_line(ab, zip_path)
    ab2 = read_ab_line(zip_path)

    assert abs(ab2.a_lon - ab.a_lon) < 1e-7, f"a_lon drift: {abs(ab2.a_lon - ab.a_lon)}"
    assert abs(ab2.a_lat - ab.a_lat) < 1e-7
    assert abs(ab2.b_lon - ab.b_lon) < 1e-7
    assert abs(ab2.b_lat - ab.b_lat) < 1e-7


def test_ab_line_roundtrip_zip_contains_shp(tmp_path):
    ab = ABLine(a_lon=-89.10, a_lat=42.56, b_lon=-89.11, b_lat=42.57)
    zip_path = tmp_path / "ab.zip"
    write_ab_line(ab, zip_path)
    with zipfile.ZipFile(zip_path) as z:
        names = z.namelist()
    assert any(n.endswith(".shp") for n in names)
    assert any(n.endswith(".prj") for n in names)


# ---------------------------------------------------------------------------
# KML
# ---------------------------------------------------------------------------


def test_kml_created(tmp_path, numeric_plots, numeric_trial):
    out = write_kml(numeric_plots, numeric_trial, "synthetic", tmp_path)
    assert out.exists()
    assert out.suffix == ".kml"


def test_kml_is_valid_xml(tmp_path, numeric_plots, numeric_trial):
    import xml.etree.ElementTree as ET
    out = write_kml(numeric_plots, numeric_trial, "synthetic", tmp_path)
    ET.parse(out)  # raises if not valid XML


def test_kml_has_one_placemark_per_plot(tmp_path, numeric_plots, numeric_trial):
    import xml.etree.ElementTree as ET
    out = write_kml(numeric_plots, numeric_trial, "synthetic", tmp_path)
    tree = ET.parse(out)
    ns = {"kml": "http://www.opengis.net/kml/2.2"}
    placemarks = tree.findall(".//kml:Placemark", ns)
    assert len(placemarks) == len(numeric_plots)


def test_kml_categorical(tmp_path, categorical_plots, categorical_trial):
    out = write_kml(categorical_plots, categorical_trial, "synthetic", tmp_path)
    content = out.read_text()
    assert "Tank Mix A" in content or "TankMixA" in content


# ---------------------------------------------------------------------------
# Sample pins XLSX
# ---------------------------------------------------------------------------


def test_sample_pins_created(tmp_path, numeric_plots):
    out = write_sample_pins(numeric_plots, "synthetic", tmp_path)
    assert out.exists()
    assert out.suffix == ".xlsx"


def test_sample_pins_one_row_per_block(tmp_path, numeric_plots):
    out = write_sample_pins(numeric_plots, "synthetic", tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    n_reps = len(set(p.rep for p in numeric_plots))
    # Row 1 = header; rows 2..n_reps+1 = data
    data_rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0] is not None and str(row[0]).startswith("Block")
    ]
    assert len(data_rows) == n_reps


def test_sample_pins_has_latlon(tmp_path, numeric_plots):
    out = write_sample_pins(numeric_plots, "synthetic", tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    lat = ws.cell(row=2, column=3).value
    lon = ws.cell(row=2, column=4).value
    assert lat is not None and isinstance(lat, float)
    assert lon is not None and isinstance(lon, float)
    assert 40 < lat < 50  # roughly Midwest latitude
    assert -100 < lon < -80  # roughly Midwest longitude


def test_sample_pins_has_hyperlinks(tmp_path, numeric_plots):
    out = write_sample_pins(numeric_plots, "synthetic", tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    gmaps = ws.cell(row=2, column=5)
    amaps = ws.cell(row=2, column=6)
    assert gmaps.hyperlink is not None
    assert amaps.hyperlink is not None
    assert "google.com" in gmaps.hyperlink.target
    assert "apple.com" in amaps.hyperlink.target


# ---------------------------------------------------------------------------
# Summary markdown
# ---------------------------------------------------------------------------


def test_summary_created(tmp_path, numeric_trial, numeric_plots, soil_ok):
    out = write_summary(numeric_trial, numeric_plots, _BLOCKS_2REP, soil_ok, _AB, "synthetic", tmp_path)
    assert out.exists()
    assert out.suffix == ".md"


def test_summary_contains_trial_type(tmp_path, numeric_trial, numeric_plots, soil_ok):
    out = write_summary(numeric_trial, numeric_plots, _BLOCKS_2REP, soil_ok, _AB, "synthetic", tmp_path)
    content = out.read_text()
    assert "fertility" in content.lower()


def test_summary_contains_rep_count(tmp_path, numeric_trial, numeric_plots, soil_ok):
    out = write_summary(numeric_trial, numeric_plots, _BLOCKS_2REP, soil_ok, _AB, "synthetic", tmp_path)
    content = out.read_text()
    assert "2" in content  # 2 reps


def test_summary_soil_unavailable_note(tmp_path, numeric_trial, numeric_plots, soil_unavailable):
    out = write_summary(numeric_trial, numeric_plots, _BLOCKS_2REP, soil_unavailable, _AB, "synthetic", tmp_path)
    content = out.read_text()
    assert "unavailable" in content.lower() or "Timed out" in content


def test_summary_categorical_note(tmp_path, categorical_trial, categorical_plots, soil_ok):
    out = write_summary(categorical_trial, categorical_plots, _BLOCKS_2REP, soil_ok, _AB, "synthetic", tmp_path)
    content = out.read_text()
    assert "categorical" in content.lower() or "No FieldView" in content


def test_summary_no_ab_line(tmp_path, numeric_trial, numeric_plots, soil_ok):
    out = write_summary(numeric_trial, numeric_plots, _BLOCKS_2REP, soil_ok, None, "synthetic", tmp_path)
    content = out.read_text()
    assert "No AB line" in content


# ---------------------------------------------------------------------------
# Map render (PNG + PDF)
# ---------------------------------------------------------------------------


def test_map_renders_png_and_pdf(tmp_path, numeric_plots, numeric_trial):
    png, pdf = write_map(numeric_plots, numeric_trial, "synthetic", tmp_path)
    assert png.exists() and png.suffix == ".png"
    assert pdf.exists() and pdf.suffix == ".pdf"


def test_map_png_is_nonzero(tmp_path, numeric_plots, numeric_trial):
    png, _ = write_map(numeric_plots, numeric_trial, "synthetic", tmp_path)
    assert png.stat().st_size > 10_000  # at least 10 KB


def test_map_categorical_trial(tmp_path, categorical_plots, categorical_trial):
    png, pdf = write_map(categorical_plots, categorical_trial, "synthetic", tmp_path)
    assert png.exists()
    assert pdf.exists()
