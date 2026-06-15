"""Write the plot-flagging XLSX with clickable Google Maps and Apple Maps hyperlinks."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from trialfield_core.geometry.placement import RepBlock
from trialfield_core.geometry.uv_frame import UVFrame

_LINK_FONT = Font(color="0000FF", underline="single")
_HEADER_FONT = Font(bold=True)
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")


def _nav_urls(lat: float, lon: float, label: str) -> tuple[str, str]:
    lat_r, lon_r = round(lat, 7), round(lon, 7)
    gmaps = f"https://www.google.com/maps?q={lat_r},{lon_r}"
    amaps = f"https://maps.apple.com/?ll={lat_r},{lon_r}&q={label.replace(' ', '%20')}"
    return gmaps, amaps


def _append_pin_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label: str,
    lat: float,
    lon: float,
    gmaps_url: str,
    amaps_url: str,
) -> None:
    lat_r, lon_r = round(lat, 7), round(lon, 7)
    ws.append([label, lat_r, lon_r, "Google Maps", "Apple Maps", f"{lat_r}, {lon_r}"])
    row = ws.max_row

    gmaps_cell = ws.cell(row=row, column=4)
    gmaps_cell.value = "Open in Google Maps"
    gmaps_cell.hyperlink = gmaps_url
    gmaps_cell.font = _LINK_FONT

    amaps_cell = ws.cell(row=row, column=5)
    amaps_cell.value = "Open in Apple Maps"
    amaps_cell.hyperlink = amaps_url
    amaps_cell.font = _LINK_FONT


def _append_section_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title: str,
) -> None:
    ws.append([])
    ws.append([title])
    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = _HEADER_FONT
    title_cell.fill = _SECTION_FILL

    ws.append(["Point", "Latitude", "Longitude", "Google Maps", "Apple Maps", "Coordinates"])
    for col in range(1, 7):
        ws.cell(row=ws.max_row, column=col).font = _HEADER_FONT


def write_flagging_pins(
    blocks: list[RepBlock],
    frame: UVFrame,
    trial_name: str,
    out_dir: Path,
) -> Path:
    """Write flagging-pin XLSX with exterior trial corners and all rep corners."""
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{trial_name}_flagging_pins.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flagging Pins"

    # --- Trial Exterior Corners ---
    u_min = min(b.u_west for b in blocks)
    u_max = max(b.u_east for b in blocks)
    v_min = min(b.v_south for b in blocks)
    v_max = max(b.v_north for b in blocks)

    exterior_corners = [
        ("Trial SW Corner", u_min, v_min),
        ("Trial SE Corner", u_max, v_min),
        ("Trial NE Corner", u_max, v_max),
        ("Trial NW Corner", u_min, v_max),
    ]

    _append_section_header(ws, "TRIAL EXTERIOR CORNERS")
    for label, u, v in exterior_corners:
        lon, lat = frame.uv_to_wgs84(u, v)
        gmaps, amaps = _nav_urls(lat, lon, label)
        _append_pin_row(ws, label, lat, lon, gmaps, amaps)

    # --- Replication Corners ---
    _append_section_header(ws, "REPLICATION CORNERS")
    for block in sorted(blocks, key=lambda b: b.rep):
        rep_corners = [
            (f"Rep {block.rep} ({block.label}) SW", block.u_west, block.v_south),
            (f"Rep {block.rep} ({block.label}) SE", block.u_east, block.v_south),
            (f"Rep {block.rep} ({block.label}) NE", block.u_east, block.v_north),
            (f"Rep {block.rep} ({block.label}) NW", block.u_west, block.v_north),
        ]
        for label, u, v in rep_corners:
            lon, lat = frame.uv_to_wgs84(u, v)
            gmaps, amaps = _nav_urls(lat, lon, label)
            _append_pin_row(ws, label, lat, lon, gmaps, amaps)

    # Notes
    ws.append([])
    ws.append(["Notes"])
    ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT
    for note in [
        "• Tap Google Maps or Apple Maps to navigate directly to each flagging point.",
        "• Flag Trial Exterior Corners first to mark the overall trial boundary.",
        "• Then flag Replication Corners to define each rep block.",
    ]:
        ws.append([note])

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    wb.save(out_path)
    return out_path
