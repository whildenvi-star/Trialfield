"""Write the soil-sample-pin XLSX with Google Maps and Apple Maps clickable hyperlinks."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from trialfield_core.geometry.plots import PlotRecord


def _block_center(plots: list[PlotRecord]) -> tuple[float, float]:
    """Return (lat, lon) center of a block from its constituent PlotRecords."""
    all_lats = [p.sw_lat for p in plots] + [p.ne_lat for p in plots]
    all_lons = [p.sw_lon for p in plots] + [p.ne_lon for p in plots]
    return (min(all_lats) + max(all_lats)) / 2, (min(all_lons) + max(all_lons)) / 2


def write_sample_pins(
    plots: list[PlotRecord],
    trial_name: str,
    out_dir: Path,
) -> Path:
    """Write sample-pin XLSX with one row per rep block; return the output path."""
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{trial_name}_soil_sample_pins.xlsx"

    # Group plots by rep, preserving rep order
    by_rep: dict[int, list[PlotRecord]] = {}
    for p in plots:
        by_rep.setdefault(p.rep, []).append(p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Pins"

    # Header row
    headers = [
        "Block", "Position", "Latitude", "Longitude",
        "Google Maps Link", "Apple Maps Link", "Coordinates (copy/paste)",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Data rows
    for rep_num in sorted(by_rep):
        rep_plots = by_rep[rep_num]
        position = rep_plots[0].rep_position
        lat, lon = _block_center(rep_plots)
        lat_r = round(lat, 7)
        lon_r = round(lon, 7)

        block_label = f"Block {rep_num}"
        gmaps_url = f"https://www.google.com/maps?q={lat_r},{lon_r}"
        amaps_url = f"https://maps.apple.com/?ll={lat_r},{lon_r}&q={block_label.replace(' ', '%20')}"
        coord_str = f"{lat_r}, {lon_r}"

        row_idx = ws.max_row + 1
        ws.append([block_label, position, lat_r, lon_r, "Open in Google Maps", "Open in Apple Maps", coord_str])

        # Add hyperlinks to Google and Apple Maps cells
        gmaps_cell = ws.cell(row=row_idx, column=5)
        gmaps_cell.value = f"Open {block_label} in Google Maps"
        gmaps_cell.hyperlink = gmaps_url
        gmaps_cell.font = Font(color="0000FF", underline="single")

        amaps_cell = ws.cell(row=row_idx, column=6)
        amaps_cell.value = f"Open {block_label} in Apple Maps"
        amaps_cell.hyperlink = amaps_url
        amaps_cell.font = Font(color="0000FF", underline="single")

    # Notes section
    ws.append([])
    ws.append(["Notes"])
    ws["A" + str(ws.max_row)].font = Font(bold=True)
    notes = [
        "• Pins are at the geographic center of each block.",
        "• Tap a Google Maps or Apple Maps link to open turn-by-turn navigation on phone.",
        "• Pull a composite soil sample within ~30 ft radius of each pin.",
        "• Sampling depth 0–8 inches recommended for baseline N/OM/CEC/pH.",
    ]
    for note in notes:
        ws.append([note])

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 24

    wb.save(out_path)
    return out_path
